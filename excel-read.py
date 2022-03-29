import openpyxl

# note: open workbook
wb = openpyxl.load_workbook('Doc-12-17 v1.xlsx')

# note: get all the sheet names
# print(wb.sheetnames)

# note: get active sheet - if sheet is open then where the cursor is, if sheet is close
# then the sheet which was last open before closing the sheet
# print(wb.active)

# note: get sheet object
# sheetNames = wb.sheetnames
# print(wb[sheetNames[0]])

# if name is already known
# print(wb['first'])

# note: get sheet title as string
firstSheet = wb['first']
# print(firstSheet.title)

# note: get cell
# print(firstSheet['a1'])

# note: get cell value
cellA1 = firstSheet['a1']
# print(cellA1.value)

# note: get cell row, column
cell_a1_row = firstSheet['b1'].row
cell_a1_column = firstSheet['a1'].column
print(cell_a1_row)
print(cell_a1_column)

# note: range values in tuple
area_cells = firstSheet['a1':'c4']
print(area_cells)
































